"""상담매뉴얼 원본 .xls → 개인정보 블라인드 사본(.xlsx) — 로컬 실행 도구(배포 제외).

원본 양식을 유지한 채(시트·셀 배치·병합·열너비·행높이·줄바꿈) 텍스트 셀에만
블라인드 규칙을 적용해 파일별 사본을 만든다. 원본은 절대 수정하지 않는다.

블라인드 규칙(적용 순서):
  1. mask_pii (parse_counsel_xls 재사용 — 카드·계좌·식별번호, 산식 오마스킹 가드 포함)
  2. 전화번호 — 휴대(01X)·유선/내선(0XX) → 0**-****-**** (자리수 무관 고정 토큰)
  3. 이메일(완전형 도메인만) → ***@***  ("P@ssw0rd1" 같은 예시 문자열 오탐 방지)
  4. 실명 — 성만 남기고 마스킹(홍길동→홍**):
     · 이름+직함님: "홍길동 차장님", "김철수과장님" (님 접미 = 인명 강신호)
     · 이름+대리: "성춘향 대리" — (?!인) 부정선행으로 도메인 용어
       '대리인'(상임/주문/개설대리인…)은 절대 건드리지 않음
     · 담당자 표기: "담당자: 홍길동", "담당자 (홍길동 대리)"
     직함 앞 역할 접두(업무/부서/지점/상임/해당…)는 인명이 아니므로 제외.

한계(요약 파일에 명시): 직함·담당자 표기 없이 이름만 단독 등장하는 실명은
사전 없이 탐지 불가 — 규칙 밖 잔존 가능. 고객사명은 기존 결정(보존)을 따른다.

출력: presentations/상담매뉴얼_블라인드/{원본명}.xlsx + _블라인드_처리내역.md
(.xls 쓰기는 미지원 형식이라 확장자만 .xlsx로 변환, 배치는 동일)

  .venv/bin/python deploy/online/blind_counsel_xls.py            # 전량 변환
  .venv/bin/python deploy/online/blind_counsel_xls.py --dry-run  # 마스킹 건수만 집계
"""
from __future__ import annotations

import argparse
import datetime as _dt
import pathlib
import re
import sys

import xlrd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

HERE = pathlib.Path(__file__).resolve().parent
ROOT = HERE.parents[1]
sys.path.insert(0, str(HERE))
from parse_counsel_xls import mask_pii  # noqa: E402 — 카드·계좌·식별번호(수정 금지)

SRC_DIR = ROOT / "코스콤(주)PB고객지원센터_ 상담매뉴얼"
OUT_DIR = ROOT / "presentations" / "상담매뉴얼_블라인드"
SUMMARY = OUT_DIR / "_블라인드_처리내역.md"

# ── 블라인드 규칙 ───────────────────────────────────────────────────────────
RE_PHONE = re.compile(r"(?<!\d)01[016789][-.\s]?\d{3,4}[-.\s]?\d{4}(?!\d)"
                      r"|(?<!\d)0\d{1,2}[-.)\s]\s?\d{3,4}[-.\s]\d{4}(?!\d)")
RE_EMAIL = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9-]+\.[A-Za-z]{2,}")

_TITLES_NIM = r"(?:차장|과장|부장|팀장|사원|주임|매니저|대리)"
# 이름+직함님 — '님' 접미가 인명 강신호. 이름 앞은 한글이 아니어야(합성어 방지).
RE_NAME_TITLE_NIM = re.compile(rf"(?<![가-힣])([가-힣]{{2,3}})\s?({_TITLES_NIM})(님)")
# 이름+대리 — 도메인 용어 '대리인'은 (?!인)으로 제외.
RE_NAME_DAERI = re.compile(r"(?<![가-힣])([가-힣]{2,3})\s(대리)(?!인)(?![가-힣])")
# 담당자 표기: "담당자: 홍길동", "담당자 (홍길동 대리)"
RE_DAMDANG = re.compile(r"(담당자\s*[:：(]\s*)([가-힣]{2,3})(\s?(?:대리|과장|차장|부장|팀장|주임)?)")
# 직함 없는 이름+님 — 연락 문맥 접미(께/에게/한테)로 한정("홍길동님께 문의").
RE_NAME_NIM = re.compile(r"(?<![가-힣])([가-힣]{3})(님(?:께|에게|한테))")

# 직함 앞에 오지만 인명이 아닌 역할·수식 접두(오마스킹 방지)
_ROLE_STOP = {
    "업무", "부서", "지점", "본사", "본부", "센터", "해당", "담당", "당사",
    "상임", "개설", "주문", "계좌", "법인", "온라인", "그쪽", "저희", "우리",
    "공통", "관리자", "책임자", "담당자", "고객", "회원사",
}


def _is_role(name: str) -> bool:
    """인명이 아닌 역할·수식어 판정(완전일치 + 2자 접두: '공통파트'→'공통')."""
    return name in _ROLE_STOP or name[:2] in _ROLE_STOP


def _mask_name(name: str) -> str:
    return name[0] + "*" * (len(name) - 1) if len(name) > 1 else name


def blind_text(t: str) -> tuple[str, int]:
    """텍스트 1셀에 블라인드 규칙 적용. 반환: (블라인드 텍스트, 이벤트 수)."""
    n = 0
    before = t
    t = mask_pii(t)
    if t != before:
        n += 1

    def phone(_m):
        nonlocal n
        n += 1
        return "0**-****-****"

    def email(_m):
        nonlocal n
        n += 1
        return "***@***"

    def name_title(m):
        nonlocal n
        name = m.group(1)
        if _is_role(name):
            return m.group(0)
        n += 1
        return f"{_mask_name(name)} {m.group(2)}{m.group(3) if m.lastindex >= 3 and m.group(3) else ''}"

    def damdang(m):
        nonlocal n
        name = m.group(2)
        if _is_role(name):
            return m.group(0)
        n += 1
        return f"{m.group(1)}{_mask_name(name)}{m.group(3)}"

    def name_nim(m):
        nonlocal n
        name = m.group(1)
        if _is_role(name):
            return m.group(0)
        n += 1
        return f"{_mask_name(name)}{m.group(2)}"

    t = RE_PHONE.sub(phone, t)
    t = RE_EMAIL.sub(email, t)
    t = RE_NAME_TITLE_NIM.sub(name_title, t)
    t = RE_NAME_DAERI.sub(name_title, t)
    t = RE_DAMDANG.sub(damdang, t)
    t = RE_NAME_NIM.sub(name_nim, t)
    return t, n


# ── 양식 보존 변환 ──────────────────────────────────────────────────────────
def convert(src: pathlib.Path, out: pathlib.Path) -> dict:
    """원본 .xls 1개 → 블라인드 .xlsx. 반환: {sheets, cells, masked}."""
    book = xlrd.open_workbook(str(src), formatting_info=True)
    wb = Workbook()
    wb.remove(wb.active)
    stat = {"sheets": 0, "cells": 0, "masked": 0}

    for sh in book.sheets():
        ws = wb.create_sheet(title=sh.name[:31])
        stat["sheets"] += 1
        # 열너비(1/256문자 → 문자폭) · 행높이(twip → pt)
        for c, info in sh.colinfo_map.items():
            if info.width:
                ws.column_dimensions[get_column_letter(c + 1)].width = round(info.width / 256, 2)
        for r, info in sh.rowinfo_map.items():
            if info.height:
                ws.row_dimensions[r + 1].height = round(info.height / 20, 1)
        # 셀 값 + 블라인드
        wrap = Alignment(wrap_text=True, vertical="top")
        bold = Font(bold=True)
        for r in range(sh.nrows):
            for c in range(sh.ncols):
                ctype = sh.cell_type(r, c)
                if ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    continue
                v = sh.cell_value(r, c)
                if ctype == xlrd.XL_CELL_TEXT:
                    v, k = blind_text(v)
                    stat["masked"] += k
                elif ctype == xlrd.XL_CELL_DATE:
                    try:
                        v = xlrd.xldate.xldate_as_datetime(v, book.datemode)
                    except Exception:  # noqa: BLE001 — 비정상 날짜는 원값 유지
                        pass
                elif ctype == xlrd.XL_CELL_BOOLEAN:
                    v = bool(v)
                elif ctype == xlrd.XL_CELL_ERROR:
                    v = None
                cell = ws.cell(row=r + 1, column=c + 1, value=v)
                cell.alignment = wrap
                if r == 0:
                    cell.font = bold
                if isinstance(v, _dt.datetime):
                    cell.number_format = "yyyy-mm-dd"
                stat["cells"] += 1
        # 병합 범위
        for rlo, rhi, clo, chi in sh.merged_cells:
            if rhi - rlo > 1 or chi - clo > 1:
                ws.merge_cells(start_row=rlo + 1, end_row=rhi,
                               start_column=clo + 1, end_column=chi)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return stat


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="상담매뉴얼 .xls 블라인드 사본 생성")
    ap.add_argument("--src", type=pathlib.Path, default=SRC_DIR)
    ap.add_argument("--out", type=pathlib.Path, default=OUT_DIR)
    ap.add_argument("--dry-run", action="store_true", help="파일 생성 없이 마스킹 건수만")
    args = ap.parse_args(argv)

    files = sorted(args.src.glob("*.xls"))
    if not files:
        raise SystemExit(f"[blind] 원본 .xls 없음: {args.src}")

    rows = []
    for f in files:
        if args.dry_run:
            book = xlrd.open_workbook(str(f))
            masked = cells = 0
            for sh in book.sheets():
                for r in range(sh.nrows):
                    for c in range(sh.ncols):
                        v = sh.cell_value(r, c)
                        if isinstance(v, str) and v:
                            cells += 1
                            masked += blind_text(v)[1]
            rows.append((f.name, "-", cells, masked))
        else:
            out = args.out / (f.stem + ".xlsx")
            st = convert(f, out)
            rows.append((f.name, st["sheets"], st["cells"], st["masked"]))
        print(f"[blind] {f.name} → 마스킹 {rows[-1][3]}건", flush=True)

    total = sum(r[3] for r in rows)
    if not args.dry_run:
        lines = [
            "# 상담매뉴얼 블라인드 처리 내역",
            "",
            f"생성: {_dt.date.today().isoformat()} · 도구: `deploy/online/blind_counsel_xls.py` · 원본 무수정",
            "",
            "적용 규칙: ① 카드·계좌·식별번호(mask_pii 재사용) ② 전화번호(휴대·유선·내선) "
            "③ 이메일(완전형만) ④ 실명 — 이름+직함님/이름+대리(도메인 용어 '대리인' 제외)/담당자 표기, "
            "성만 남김(홍길동→홍**). 고객사명은 기존 결정대로 보존.",
            "",
            "한계: 직함·담당자 표기 없이 단독 등장하는 이름은 규칙으로 탐지 불가(잔존 가능). "
            "발견 시 규칙 추가 후 재실행(원본 불변이므로 반복 가능).",
            "",
            "| 파일 | 시트 | 텍스트 셀 | 마스킹 이벤트 |",
            "|---|---|---|---|",
        ]
        lines += [f"| {n} | {s} | {c:,} | {m:,} |" for n, s, c, m in rows]
        lines += ["", f"**합계: 파일 {len(rows)}개 · 마스킹 {total:,}건** (확장자만 .xls→.xlsx, 배치 동일)"]
        SUMMARY.parent.mkdir(parents=True, exist_ok=True)
        (args.out / SUMMARY.name).write_text("\n".join(lines) + "\n", encoding="utf-8")
        print(f"[blind] 요약 → {args.out / SUMMARY.name}")
    print(f"[blind] 완료 — 파일 {len(rows)}개, 마스킹 {total:,}건")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
