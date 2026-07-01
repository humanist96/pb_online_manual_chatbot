"""
구조화 트리(parse.py) → 샘플과 동일 계열의 XLSX (B/C 2열 레이아웃).

메타 블록은 HTML/URL에서 유도 가능한 값만 채우고(제목·코드·유형·파일명·화면TR),
발행부서·버전·작성일·성명·일자는 공란(원본 .doc 문서정보에 존재, HTML엔 없음).

사용:
  python src/to_xlsx.py data/html/AC250400.html            # -> data/xlsx/*.xlsx
  python src/to_xlsx.py data/html/AC250400.html -o out.xlsx
"""
from __future__ import annotations
import sys
import glob
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

import re

from parse import parse_html


def safe_filename(name: str) -> str:
    """파일명에 쓸 수 없는 문자 치환 (/ \\ : * ? " < > |)."""
    return re.sub(r'[\\/:*?"<>|]+', "_", name).strip()

HDR_FILL = PatternFill("solid", fgColor="DCE6F1")
SECT_FILL = PatternFill("solid", fgColor="F2F2F2")
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")


def build_workbook(doc: dict) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = doc["screen_id"]

    r = 1

    def put(b=None, c=None, *, bold=False, fill=None, span=False):
        nonlocal r
        if b is not None:
            cell = ws.cell(row=r, column=2, value=b)
            if bold:
                cell.font = BOLD
            if fill:
                cell.fill = fill
            cell.alignment = WRAP_TOP
        if c is not None:
            cell = ws.cell(row=r, column=3, value=c)
            cell.alignment = WRAP_TOP
            if fill:
                cell.fill = fill
        r += 1

    # ── 메타정보 블록 ──
    put("<메타정보>", bold=True, fill=HDR_FILL)
    filename = f"{doc['screen_no']}_{doc['screen_id']}_{doc['title']}.doc"
    tr = f"TR-{doc['screen_no']}" if doc["screen_no"] else ""
    meta = [
        ("제목", doc["title"]),
        ("코드", doc["code"]),
        ("유형", "온라인매뉴얼 문서 정보"),
        ("발행 부서", ""),          # HTML 미포함 → 공란
        ("파일명", filename),
        ("버전", ""),               # HTML 미포함 → 공란
        ("작성일", ""),             # HTML 미포함 → 공란
        ("화면TR", tr),
        ("성명", ""),               # HTML 미포함 → 공란
        ("일자", ""),               # HTML 미포함 → 공란
    ]
    for k, v in meta:
        put(k, v, bold=True)

    r += 1  # 빈 줄

    # ── 본문추출 ──
    put("<본문추출>", bold=True, fill=HDR_FILL)
    put(f"AUP: {doc['aup']}", bold=True)
    put("요약", doc["summary"], bold=True, fill=SECT_FILL)
    r += 1

    # ── 브레드크럼 (경로 | 설명) ──
    put("경로", "설명", bold=True, fill=HDR_FILL)
    for bc in doc["breadcrumbs"]:
        put(" > ".join(bc["path"]), bc["text"])

    # 열 폭/줄바꿈
    ws.column_dimensions[get_column_letter(1)].width = 2
    ws.column_dimensions[get_column_letter(2)].width = 55
    ws.column_dimensions[get_column_letter(3)].width = 80
    ws.sheet_view.showGridLines = True
    return wb


def main():
    args = sys.argv[1:]
    out = None
    files = []
    it = iter(args)
    for a in it:
        if a in ("-o", "--out"):
            out = next(it)
        else:
            files.extend(glob.glob(a))
    if not files:
        print("usage: python src/to_xlsx.py <html...> [-o out.xlsx]", file=sys.stderr)
        sys.exit(1)

    outdir = pathlib.Path("data/xlsx")
    outdir.mkdir(parents=True, exist_ok=True)
    for f in sorted(files):
        doc = parse_html(f)
        wb = build_workbook(doc)
        if out and len(files) == 1:
            target = out
        else:
            target = str(outdir / safe_filename(
                f"{doc['screen_no']}_{doc['screen_id']}_{doc['title']}.xlsx"))
        try:
            wb.save(target)
            print(f"wrote {target}  (rows={len(doc['breadcrumbs'])+15})", file=sys.stderr)
        except PermissionError:
            # 대상 파일이 Excel 등에서 열려 있어 잠김 → 건너뛰고 계속
            print(f"SKIP  {target}  (파일이 열려 있어 잠김 — Excel 등을 닫고 재실행)", file=sys.stderr)


if __name__ == "__main__":
    main()
